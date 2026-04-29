from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from django.utils import timezone
from system_app.models import Account, RawMaterial
from .forms import RawMaterialForm, AddEmployeeForm, EditEmployeeNameForm
from .models import MaterialUnit, Employee, EmployeeRole
import random
from .models import Product, ProductCategory, ProductCollection, ProductMaterial, Order, OrderItem  # ✅ FIX: Added OrderItem import

from django.db import models as db_models

import json
from django.core.serializers.json import DjangoJSONEncoder
from datetime import datetime
from calendar import month_abbr
from django.urls import reverse
from urllib.parse import urlencode
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST

#Global variable for login authentication
account_id = 0

# Function for login page
def login_view(request):
    message = request.session.pop('message', '')

    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        password = request.POST.get('password')

        try:
            account = Account.objects.get(employee_id=employee_id, password=password)
            if password == account.password:
                request.session['account_id'] = account.id
                request.session['employee_id'] = account.employee_id

                try:
                    from .models import Employee
                    emp = Employee.objects.get(employee_id=employee_id)
                    request.session['employee_name'] = emp.employee_name
                except Employee.DoesNotExist:
                    request.session['employee_name'] = 'Owner'

                if employee_id.startswith('0'):
                    return redirect('owner_homepage')
                elif employee_id.startswith('1'):
                    return redirect('prodman_homepage')
                elif employee_id.startswith('2'):
                    return redirect('prodemp_home')
                else:
                    message = "Unrecognized account type. Please contact your administrator."
            else:
                message = "Invalid login. Please check your username or password."
        except Account.DoesNotExist:
            message = "Invalid login. Please check your username or password."

    return render(request, 'system_app/login.html', {'message': message})

def logout_view(request):
    request.session.pop('account_id', None)
    return redirect('login')

def change_password_view(request):
    account_id = request.session.get('account_id')
    if not account_id:
        return redirect('login')

    message = ""

    if request.method == "POST":
        current_pw = request.POST.get("current_password")
        new_pw = request.POST.get("new_password")
        confirm_pw = request.POST.get("confirm_password")

        account = Account.objects.get(id=account_id)

        if current_pw != account.password:
            message = "Current password is incorrect."
        elif not new_pw:
            message = "New password cannot be empty."
        elif new_pw != confirm_pw:
            message = "New passwords do not match."
        else:
            account.password = new_pw
            account.save()
            message = "Password changed successfully!"

    return render(request, "system_app/change_password.html", {"message": message})


def owner_homepage(request):
    if not request.session.get('account_id'):
        return redirect('login')
    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('0'):
        return redirect('login')
    return render(request, 'system_app/owner_home.html', {
        'employee_id': employee_id,
    })

def owner_sales_report(request):
    if not request.session.get('account_id'):
        return redirect('login')
    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('0'):
        return redirect('login')

    orders = Order.objects.prefetch_related('items__product__product_materials__raw_material').all()
    total_sales = 0.0
    total_expenses = 0.0
    total_products_sold = 0
    product_quantities = {}

    today = timezone.now()
    month_keys = []
    month_data = {}
    for i in range(5, -1, -1):
        month = today.month - i
        year = today.year
        while month <= 0:
            month += 12
            year -= 1
        key = (year, month)
        month_keys.append(key)
        month_data[key] = {
            'label': month_abbr[month],
            'sales': 0.0,
            'expenses': 0.0,
        }

    for order in orders:
        for item in order.items.all():
            product = item.product
            qty = item.quantity
            total_products_sold += qty
            item_sales = product.price * qty
            total_sales += item_sales
            product_quantities[product.product_name] = product_quantities.get(product.product_name, 0) + qty

            item_cost = 0.0
            for pm in product.product_materials.all():
                if pm.quantity_per_garment and pm.raw_material.material_unitprice:
                    item_cost += pm.quantity_per_garment * pm.raw_material.material_unitprice * qty
            total_expenses += item_cost

            order_month = (order.created_at.year, order.created_at.month)
            if order_month in month_data:
                month_data[order_month]['sales'] += item_sales
                month_data[order_month]['expenses'] += item_cost

    net_profit = total_sales - total_expenses
    top_products = sorted(
        [{'name': name, 'quantity': qty} for name, qty in product_quantities.items()],
        key=lambda p: p['quantity'],
        reverse=True
    )[:5]

    max_value = max(
        [entry['sales'] for entry in month_data.values()] +
        [entry['expenses'] for entry in month_data.values()] + [1]
    )

    chart_data = []
    for key in month_keys:
        entry = month_data[key]
        chart_data.append({
            'label': entry['label'],
            'sales': entry['sales'],
            'expenses': entry['expenses'],
            'sales_pct': int((entry['sales'] / max_value) * 100) if max_value else 0,
            'expenses_pct': int((entry['expenses'] / max_value) * 100) if max_value else 0,
        })

    return render(request, 'system_app/owner_sales_report.html', {
        'total_sales': total_sales,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'total_products_sold': total_products_sold,
        'top_products': top_products,
        'chart_data': chart_data,
    })


def prodman_homepage(request):
    if not request.session.get('account_id'):
        return redirect('login')
    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('1'):
        return redirect('login')
    return render(request, 'system_app/prodman_home.html', {
        'employee_id': employee_id,
    })

def prodemp_homepage(request):
    if not request.session.get('account_id'):
        return redirect('login')
    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('2'):
        return redirect('login')
    products = Product.objects.all()
    materials = RawMaterial.objects.all()

    return render(request, 'system_app/prodemp_home.html', {
        'employee_id': employee_id,
        'products': products,
        'materials': materials,
    })

def prodman_matinv(request):
    if not request.session.get('account_id'):
        return redirect('login')

    q = (request.GET.get('q') or "").strip()
    category = (request.GET.get('category') or "all").strip()
    sort = (request.GET.get('sort') or "alpha").strip()

    qs = RawMaterial.objects.all()

    if category in {"fabrics", "trims", "accessories"}:
        qs = qs.filter(material_category=category)

    if q:
        qs = qs.filter(material_name__icontains=q)

    if sort == "highest":
        qs = qs.order_by("-material_quantity", "material_name")
    elif sort == "lowest":
        qs = qs.order_by("material_quantity", "material_name")
    elif sort == "alpha_desc":
        qs = qs.order_by("-material_name")
    else:
        qs = qs.order_by("material_name")

    if request.method == "POST":
        form = RawMaterialForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.full_clean()
            obj.save()
            params = {}
            if category:
                params["category"] = category
            if q:
                params["q"] = q
            url = reverse("prodman_matinv")
            if params:
                url = f"{url}?{urlencode(params)}"
            return redirect(url)
    else:
        form = RawMaterialForm()

    units_by_category = {
        "fabrics": [
            {"id": u.id, "name": u.unit_name}
            for u in MaterialUnit.objects.filter(category="fabrics").order_by("unit_name")
        ],
        "trims": [
            {"id": u.id, "name": u.unit_name}
            for u in MaterialUnit.objects.filter(category="trims").order_by("unit_name")
        ],
        "accessories": [
            {"id": u.id, "name": u.unit_name}
            for u in MaterialUnit.objects.filter(category="accessories").order_by("unit_name")
        ],
    }

    edit_id = request.GET.get("edit")
    edit_target = None
    edit_form = None

    if edit_id:
        try:
            edit_target = RawMaterial.objects.get(pk=int(edit_id))
            edit_form = RawMaterialForm(instance=edit_target)
        except (RawMaterial.DoesNotExist, ValueError):
            edit_target = None
            edit_form = None

    # Critical materials — quantity at or below minimum threshold
    critical_materials = RawMaterial.objects.filter(
        material_quantity__lte=db_models.F('minimum_threshold'),
        minimum_threshold__gt=0
    ).order_by('material_name')

    context = {
        "materials": qs,
        "selected_category": category,
        "q": q,
        "sort": sort,
        "as_of": timezone.localtime(timezone.now()),
        "category_choices": [
            ("all", "All"),
            ("fabrics", "Fabrics"),
            ("trims", "Trims"),
            ("accessories", "Accessories"),
        ],
        "form": form,
        "units_by_category_json": json.dumps(units_by_category, cls=DjangoJSONEncoder),
        "edit_target": edit_target,
        "edit_form": edit_form,
        "open_edit": bool(edit_target),
        "critical_materials": critical_materials,
        "critical_count": critical_materials.count(),
    }

    return render(request, "system_app/prodman_matinv.html", context)


def prod_matinv(request):
    if not request.session.get('account_id'):
        return redirect('login')

    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('2'):
        return redirect('login')

    q = (request.GET.get('q') or "").strip()
    category = (request.GET.get('category') or "all").strip()
    sort = (request.GET.get('sort') or "alpha").strip()

    qs = RawMaterial.objects.all()

    if category in {"fabrics", "trims", "accessories"}:
        qs = qs.filter(material_category=category)

    if q:
        qs = qs.filter(material_name__icontains=q)

    if sort == "highest":
        qs = qs.order_by("-material_quantity", "material_name")
    elif sort == "lowest":
        qs = qs.order_by("material_quantity", "material_name")
    elif sort == "alpha_desc":
        qs = qs.order_by("-material_name")
    else:
        qs = qs.order_by("material_name")

    units_by_category = {
        "fabrics": [
            {"id": u.id, "name": u.unit_name}
            for u in MaterialUnit.objects.filter(category="fabrics").order_by("unit_name")
        ],
        "trims": [
            {"id": u.id, "name": u.unit_name}
            for u in MaterialUnit.objects.filter(category="trims").order_by("unit_name")
        ],
        "accessories": [
            {"id": u.id, "name": u.unit_name}
            for u in MaterialUnit.objects.filter(category="accessories").order_by("unit_name")
        ],
    }

    edit_id = request.GET.get("edit")
    edit_target = None
    edit_form = None

    if edit_id:
        try:
            edit_target = RawMaterial.objects.get(pk=int(edit_id))
            edit_form = RawMaterialForm(instance=edit_target)
        except (RawMaterial.DoesNotExist, ValueError):
            edit_target = None
            edit_form = None

    context = {
        "materials": qs,
        "selected_category": category,
        "q": q,
        "sort": sort,
        "as_of": timezone.localtime(timezone.now()),
        "category_choices": [
            ("all", "All"),
            ("fabrics", "Fabrics"),
            ("trims", "Trims"),
            ("accessories", "Accessories"),
        ],
        "units_by_category_json": json.dumps(units_by_category, cls=DjangoJSONEncoder),
        "edit_target": edit_target,
        "edit_form": edit_form,
        "open_edit": bool(edit_target),
    }

    return render(request, "system_app/prod_matinv.html", context)


def edit_raw_material(request, pk):
    if not request.session.get('account_id'):
        return redirect('login')

    material = get_object_or_404(RawMaterial, pk=pk)

    q = (request.GET.get("q") or "").strip()
    category = (request.GET.get("category") or "all").strip()
    sort = (request.GET.get("sort") or "alpha").strip()

    if request.method == "POST":
        form = RawMaterialForm(request.POST, instance=material)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.full_clean()
            obj.save()

            params = {}
            if category:
                params["category"] = category
            if q:
                params["q"] = q
            if sort:
                params["sort"] = sort

            url = reverse("prodman_matinv")
            if params:
                url = f"{url}?{urlencode(params)}"
            return redirect(url)

    params = {"edit": pk}
    if category:
        params["category"] = category
    if q:
        params["q"] = q
    if sort:
        params["sort"] = sort

    url = reverse("prodman_matinv")
    return redirect(f"{url}?{urlencode(params)}")

def delete_raw_material(request, pk):
    if not request.session.get('account_id'):
        return redirect('login')

    material = get_object_or_404(RawMaterial, pk=pk)

    q = (request.GET.get("q") or "").strip()
    category = (request.GET.get("category") or "all").strip()
    sort = (request.GET.get("sort") or "alpha").strip()

    if request.method == "POST":
        material.delete()

    params = {}
    if category:
        params["category"] = category
    if q:
        params["q"] = q
    if sort:
        params["sort"] = sort

    url = reverse("prodman_matinv")
    if params:
        url = f"{url}?{urlencode(params)}"

    return redirect(url)

def _generate_employee_id(role):
    prefix = '1' if role == 'production_manager' else '2'
    existing_ids = set(Employee.objects.values_list('employee_id', flat=True))
    for _ in range(1000):
        suffix = str(random.randint(0, 999)).zfill(3)
        candidate = prefix + suffix
        if candidate not in existing_ids:
            return candidate
    raise ValueError("Could not generate a unique employee ID.")


def owner_manage_employees(request):
    if not request.session.get('account_id'):
        return redirect('login')
    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('0'):
        return redirect('login')

    message = ''
    message_type = 'success'
    add_form = AddEmployeeForm()
    edit_form = None
    edit_target = None

    if request.method == 'POST' and 'action' in request.POST:
        action = request.POST.get('action')

        if action == 'add':
            add_form = AddEmployeeForm(request.POST)
            if add_form.is_valid():
                name = add_form.cleaned_data['employee_name']
                role = add_form.cleaned_data['employee_role']
                try:
                    new_id = _generate_employee_id(role)
                    Employee.objects.create(
                        employee_id=new_id,
                        employee_name=name,
                        employee_role=role,
                    )
                    Account.objects.create(
                        employee_id=new_id,
                        password=new_id,
                    )
                    message = f"Employee added successfully! ID: {new_id} | Default password: {new_id}"
                    message_type = 'success'
                    add_form = AddEmployeeForm()
                except Exception as e:
                    message = f"Error adding employee: {e}"
                    message_type = 'danger'

        elif action == 'edit':
            edit_id = request.POST.get('edit_id')
            emp = get_object_or_404(Employee, pk=edit_id)
            edit_form = EditEmployeeNameForm(request.POST, instance=emp)
            if edit_form.is_valid():
                edit_form.save()
                message = "Employee name updated successfully."
                message_type = 'success'
            else:
                edit_target = emp
                message = "Please fix the errors below."
                message_type = 'danger'

    edit_id_get = request.GET.get('edit')
    if edit_id_get and not edit_target:
        try:
            edit_target = Employee.objects.get(pk=int(edit_id_get))
            edit_form = EditEmployeeNameForm(instance=edit_target)
        except (Employee.DoesNotExist, ValueError):
            edit_target = None

    employees = Employee.objects.all().order_by('employee_id')

    return render(request, 'system_app/owner_manage_employees.html', {
        'employees': employees,
        'add_form': add_form,
        'edit_form': edit_form,
        'edit_target': edit_target,
        'open_edit': bool(edit_target),
        'message': message,
        'message_type': message_type,
    })

def delete_employee(request, pk):
    if not request.session.get('account_id'):
        return redirect('login')
    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('0'):
        return redirect('login')

    emp = get_object_or_404(Employee, pk=pk)

    if request.method == 'POST':
        Account.objects.filter(employee_id=emp.employee_id).delete()
        emp.delete()

    return redirect('owner_manage_employees')

def owner_products(request):
    if not request.session.get('account_id'):
        return redirect('login')
    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('0'):
        return redirect('login')
    return render(request, 'system_app/owner_products.html')


from .forms import ProductForm
from .models import Product

def owner_add_product(request):
    if not request.session.get('account_id'):
        return redirect('login')
    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('0'):
        return redirect('login')

    last = Product.objects.order_by('-id').first()
    next_num = (last.id + 1) if last else 1
    next_product_id = f'#{next_num:05d}'

    message = ''
    form = ProductForm()

    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('owner_products')
        else:
            message = 'Please fix the errors below.'

    return render(request, 'system_app/owner_add_product.html', {
        'form': form,
        'next_product_id': next_product_id,
        'message': message,
    })


def owner_products_list(request):
    if not request.session.get('account_id'):
        return redirect('login')
    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('0'):
        return redirect('login')

    products = Product.objects.all().order_by('-id')
    return render(request, 'system_app/owner_products_list.html', {'products': products})

def employee_materials(request):
    if not request.session.get('account_id'):
        return redirect('login')

    materials = RawMaterial.objects.all().order_by('material_name')

    return render(request, 'system_app/prodemp_matinv.html', {
        'materials': materials
    })

def employee_products(request):
    if not request.session.get('account_id'):
        return redirect('login')

    products = Product.objects.all().order_by('product_name')

    return render(request, 'system_app/prodemp_products_list.html', {
        'products': products
    })

def prodman_products(request):

    if not request.session.get('account_id'):
        return redirect('login')

    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('1'):
        return redirect('login')

    products = Product.objects.all()

    order_items = request.session.get('order_items', [])

    if request.method == "POST":
        product_id = request.POST.get("product_id")
        qty = request.POST.get("quantity")

        if product_id and qty:
            qty = int(qty)
            product = Product.objects.get(id=product_id)

            # ✅ FIX: Only add to cart here — do NOT deduct stock yet.
            # Stock is deducted in prodman_order_summary when the order is completed.
            if qty > 0 and product.quantity >= qty:
                order_items.append({
                    'product_id': product.id,
                    'name': product.product_name,
                    'qty': qty
                })

                request.session['order_items'] = order_items
                request.session.modified = True

        return redirect('prodman_products_list')

    return render(request, 'system_app/prodman_products_list.html', {
        'products': products,
        'order_items': order_items
    })


# OWNER product detail page
def owner_product_detail(request, pk):
    if not request.session.get('account_id'):
        return redirect('login')

    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('0'):
        return redirect('login')

    product = Product.objects.get(pk=pk)
    product_materials = ProductMaterial.objects.filter(product=product).select_related('raw_material')

    # ✅ ADD THIS BLOCK (make sure spacing is EXACT)
    materials_breakdown = []
    total_cost = 0

    for pm in product_materials:
        rm = pm.raw_material
        unit_price = rm.material_unitprice or 0
        qty = pm.quantity_per_garment or 0

        subtotal = unit_price * qty
        total_cost += subtotal

        materials_breakdown.append({
            'name': rm.material_name,
            'category': rm.material_category,
            'unit': rm.material_unit.unit_name,
            'qty': qty,
            'unit_price': unit_price,
            'subtotal': subtotal,
        })

    if request.method == "POST":
        price = request.POST.get("price")
        if price:
            product.price = float(price)
            product.save()
        return redirect('owner_product_detail', pk=pk)

    return render(request, 'system_app/owner_product_detail.html', {
        'product': product,
        'product_materials': product_materials,
        'materials_breakdown': materials_breakdown,
        'total_cost': total_cost,
    })


# PRODUCTION MANAGER product detail page
def prodman_product_detail(request, pk):
    if not request.session.get('account_id'):
        return redirect('login')

    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('1'):
        return redirect('login')

    product = Product.objects.get(pk=pk)
    product_materials = list(ProductMaterial.objects.filter(product=product).select_related('raw_material', 'raw_material__material_unit'))

    if request.method == "POST":
        action = request.POST.get('action')
        if action == 'update_qty':
            new_qty = request.POST.get('quantity')
            if new_qty is not None and new_qty != '':
                try:
                    new_qty = int(new_qty)
                    qty_to_add = new_qty - product.quantity
                    if qty_to_add > 0:
                        for pm in product_materials:
                            rm = pm.raw_material
                            deduction = pm.quantity_per_garment * qty_to_add
                            rm.material_quantity = max(0, rm.material_quantity - deduction)
                            rm.save()
                    product.quantity = new_qty
                    product.save()
                except ValueError:
                    pass
        return redirect('prodman_product_detail', pk=pk)

    # Build available raw materials JSON for the Log Materials UI
    raw_materials_data = []
    for rm in RawMaterial.objects.filter(material_quantity__gt=0).order_by('material_name'):
        raw_materials_data.append({
            'id': rm.id,
            'name': rm.material_name,
            'category': rm.material_category,
            'unit': rm.material_unit.unit_name,
            'quantity': rm.material_quantity,
        })

    return render(request, 'system_app/prodman_product_detail.html', {
        'product': product,
        'product_materials': product_materials,
        'raw_materials_json': json.dumps(raw_materials_data, cls=DjangoJSONEncoder),
    })

# PRODUCTION EMPLOYEE product detail page
def prodemp_product_detail(request, pk):
    if not request.session.get('account_id'):
        return redirect('login')

    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('2'):
        return redirect('login')

    product = Product.objects.get(pk=pk)
    product_materials = list(ProductMaterial.objects.filter(product=product).select_related('raw_material', 'raw_material__material_unit'))

    if request.method == "POST":
        add_qty = request.POST.get("add_qty")
        if add_qty:
            try:
                qty_to_add = int(add_qty)
                if qty_to_add > 0:
                    product.quantity += qty_to_add
                    product.save()
                    for pm in product_materials:
                        rm = pm.raw_material
                        deduction = pm.quantity_per_garment * qty_to_add
                        rm.material_quantity = max(0, rm.material_quantity - deduction)
                        rm.save()
            except (ValueError, TypeError):
                pass
        return redirect('prodemp_product_detail', pk=pk)

    fabrics = [pm for pm in product_materials if pm.raw_material.material_category == 'fabrics']
    trims = [pm for pm in product_materials if pm.raw_material.material_category == 'trims']
    accessories = [pm for pm in product_materials if pm.raw_material.material_category == 'accessories']

    return render(request, 'system_app/prodemp_product_detail.html', {
        'product': product,
        'product_materials': product_materials,
        'fabrics': fabrics,
        'trims': trims,
        'accessories': accessories,
    })


# DELETE PRODUCT
def owner_delete_product(request, pk):
    if request.method == "POST":
        product = Product.objects.get(pk=pk)
        product.delete()

    return redirect('owner_products_list')

# EDIT PRODUCT
def owner_edit_product(request, pk):
    if not request.session.get('account_id'):
        return redirect('login')

    product = Product.objects.get(pk=pk)

    if request.method == "POST":
        product.product_name = request.POST.get("product_name")
        product.product_category = request.POST.get("product_category")
        product.product_collection = request.POST.get("product_collection")
        product.price = request.POST.get("price")
        product.save()
        return redirect('owner_product_detail', pk=product.id)

    return render(request, 'system_app/owner_edit_product.html', {
        'product': product,
        'categories': ProductCategory,
        'collections': ProductCollection
    })


# -------------------------------------------------------
# AJAX API: Log Materials for a Product
# -------------------------------------------------------

def api_raw_materials_with_stock(request):
    """Return raw materials that have stock, optionally filtered by category."""
    if not request.session.get('account_id'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    category = request.GET.get('category', '')
    qs = RawMaterial.objects.filter(material_quantity__gt=0).order_by('material_name')
    if category:
        qs = qs.filter(material_category=category)

    data = []
    for rm in qs:
        data.append({
            'id': rm.id,
            'name': rm.material_name,
            'category': rm.material_category,
            'unit': rm.material_unit.unit_name,
            'quantity': rm.material_quantity,
        })
    return JsonResponse({'materials': data})

def api_product_materials(request, pk):
    """GET: list logged materials for a product. POST: add a new material."""
    if not request.session.get('account_id'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    product = get_object_or_404(Product, pk=pk)

    if request.method == 'GET':
        pms = ProductMaterial.objects.filter(product=product).select_related('raw_material')
        data = []
        for pm in pms:
            rm = pm.raw_material
            entry = {
                'id': pm.id,
                'raw_material_id': rm.id,
                'material_name': rm.material_name,
                'category': rm.material_category,
                'unit': rm.material_unit.unit_name,
                'quantity_per_garment': pm.quantity_per_garment,
                'fabric_length': pm.fabric_length,
                'fabric_width': pm.fabric_width,
            }
            data.append(entry)
        return JsonResponse({'materials': data})

    elif request.method == 'POST':
        try:
            body = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        rm_id = body.get('raw_material_id')
        qty = body.get('quantity_per_garment')
        fabric_length = body.get('fabric_length')
        fabric_width = body.get('fabric_width')

        if not rm_id or qty is None:
            return JsonResponse({'error': 'Missing fields'}, status=400)

        try:
            qty = float(qty)
            if qty <= 0:
                raise ValueError
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid quantity'}, status=400)

        rm = get_object_or_404(RawMaterial, pk=rm_id)

        pm, created = ProductMaterial.objects.get_or_create(
            product=product,
            raw_material=rm,
            defaults={
                'quantity_per_garment': qty,
                'fabric_length': float(fabric_length) if fabric_length is not None else None,
                'fabric_width': float(fabric_width) if fabric_width is not None else None,
            }
        )

        if not created:
            pm.quantity_per_garment = qty
            pm.fabric_length = float(fabric_length) if fabric_length is not None else None
            pm.fabric_width = float(fabric_width) if fabric_width is not None else None
            pm.save()

        return JsonResponse({
            'success': True,
            'id': pm.id,
            'raw_material_id': rm.id,
            'material_name': rm.material_name,
            'category': rm.material_category,
            'unit': rm.material_unit.unit_name,
            'quantity_per_garment': pm.quantity_per_garment,
            'fabric_length': pm.fabric_length,
            'fabric_width': pm.fabric_width,
        })

    return JsonResponse({'error': 'Method not allowed'}, status=405)


def api_product_material_detail(request, pk, pm_pk):
    """PUT: update a product material. DELETE: remove it."""
    if not request.session.get('account_id'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    pm = get_object_or_404(ProductMaterial, pk=pm_pk, product__pk=pk)

    if request.method == 'PUT':
        try:
            body = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        qty = body.get('quantity_per_garment')
        fabric_length = body.get('fabric_length')
        fabric_width = body.get('fabric_width')

        if qty is None:
            return JsonResponse({'error': 'Missing quantity'}, status=400)

        try:
            qty = float(qty)
            if qty <= 0:
                raise ValueError
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid quantity'}, status=400)

        pm.quantity_per_garment = qty
        pm.fabric_length = float(fabric_length) if fabric_length is not None else None
        pm.fabric_width = float(fabric_width) if fabric_width is not None else None
        pm.save()

        return JsonResponse({
            'success': True,
            'id': pm.id,
            'quantity_per_garment': pm.quantity_per_garment,
            'fabric_length': pm.fabric_length,
            'fabric_width': pm.fabric_width,
        })

    elif request.method == 'DELETE':
        pm.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Method not allowed'}, status=405)

# order summary for prodman
def prodman_order_summary(request):

    if not request.session.get('account_id'):
        return redirect('login')

    employee_id = request.session.get('employee_id', '')
    if not employee_id.startswith('1'):
        return redirect('login')

    order_items = request.session.get('order_items', [])

    if request.method == "POST":
        action = request.POST.get('action')

        # COMPLETE ORDER
        if action == "complete":
            if order_items:
                order = Order.objects.create()
                for item in order_items:
                    try:
                        product = Product.objects.get(id=item.get('product_id'))
                        qty = item.get('qty', 0)
                        OrderItem.objects.create(
                            order=order,
                            product=product,
                            quantity=qty
                        )
                        product.quantity = max(0, product.quantity - qty)
                        product.save()
                    except Exception as e:
                        print(f"Order item error: {e}")
                        continue

            request.session['order_items'] = []
            request.session.modified = True
            return redirect('prodman_products_list')

        # UPDATE QTY
        elif action == "update":
            product_id = int(request.POST.get('product_id'))
            new_qty = int(request.POST.get('qty', 1))
            if new_qty > 0:
                for item in order_items:
                    if item['product_id'] == product_id:
                        item['qty'] = new_qty
                        break
            request.session['order_items'] = order_items
            request.session.modified = True
            return redirect('prodman_order_summary')

        # DELETE ITEM
        elif action == "delete":
            product_id = int(request.POST.get('product_id'))
            order_items = [i for i in order_items if i['product_id'] != product_id]
            request.session['order_items'] = order_items
            request.session.modified = True
            return redirect('prodman_order_summary')

    return render(request, 'system_app/prodman_order_summary.html', {
        'order_items': order_items
    })

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

def export_sales_xlsx(request):
    if not request.session.get('account_id'):
        return redirect('login')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Header row styling
    header_fill = PatternFill(start_color="C6E8ED", end_color="C6E8ED", fill_type="solid")
    header_font = Font(bold=True)

    headers = ['Order ID', 'Date', 'Product', 'Quantity', 'Unit Price (₱)', 'Subtotal (₱)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row_num = 2
    total = 0
    for item in OrderItem.objects.select_related('order', 'product').order_by('-order__created_at'):
        subtotal = round(item.quantity * item.product.price, 2)
        total += subtotal
        ws.append([
            f'#{item.order.id}',
            item.order.created_at.strftime('%Y-%m-%d %H:%M'),
            item.product.product_name,
            item.quantity,
            item.product.price,
            subtotal,
        ])
        row_num += 1

    # Total row
    ws.append([])
    total_row = row_num + 1
    ws.cell(row=total_row, column=5, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=round(total, 2)).font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    wb.save(response)
    return response


def export_sales_pdf(request):
    if not request.session.get('account_id'):
        return redirect('login')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph("<b>Two Chic Manila — Sales Report</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))

    # Table data
    data = [['Order ID', 'Date', 'Product', 'Qty', 'Unit Price', 'Subtotal']]
    total = 0

    for item in OrderItem.objects.select_related('order', 'product').order_by('-order__created_at'):
        subtotal = round(item.quantity * item.product.price, 2)
        total += subtotal
        data.append([
            f'#{item.order.id}',
            item.order.created_at.strftime('%Y-%m-%d %H:%M'),
            item.product.product_name,
            str(item.quantity),
            f'P{item.product.price:,.2f}',
            f'P{subtotal:,.2f}',
        ])

    # Total row
    data.append(['', '', '', '', 'TOTAL', f'P{round(total, 2):,.2f}'])

    table = Table(data, colWidths=[2*cm, 4*cm, 6*cm, 1.5*cm, 3*cm, 3*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C6E8ED')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#dddddd')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
    ]))
    elements.append(table)

    doc.build(elements)
    return response